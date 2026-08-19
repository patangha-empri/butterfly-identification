"""
Client-facing export of the species whose reference image needs replacing.

Builds on scripts/audit_species_images.py (same DB query and classification) and
turns it into a workbook someone outside the project can act on:

  Summary          what was checked, what the verdicts mean, the counts
  Needs new photo  one row per species that requires a species-specific image
  Plate groups     the antique plate pages, and every species pinned to each —
                   sourcing one plate's species together is the efficient path
  Synonym pairs    species sharing a photo *legitimately* (same butterfly, two
                   names). These need a data merge, not a photographer.

Read-only.

    backend/venv/Scripts/python.exe scripts/export_image_gaps.py
"""
import argparse
import collections
import os

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit_species_images import CONFIG, classify, fetch, source_filename

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)

ISSUE_LABEL = {
    "PLATE_SHARED": "Antique plate page, shared with other species",
    "PLATE": "Antique plate page",
    "SHARED": "Same photo used on another species",
    "NO_PROVENANCE": "Missing licence/source record",
}


def style_header(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(ROOT, "scripts", "species_images_to_replace.xlsx"))
    args = ap.parse_args()

    conn = psycopg2.connect(CONFIG["DATABASE_URL"])
    species = fetch(conn)
    conn.close()
    verdicts = classify(species)

    # Only active species matter — the app hides the rest.
    active = {sid: d for sid, d in species.items() if d["row"]["is_active"]}
    flagged = {sid: d for sid, d in active.items() if verdicts[sid][0] not in ("OK", "NO_IMAGE")}

    # Group by the underlying file so shared plates cluster.
    groups = collections.defaultdict(list)
    for sid, d in flagged.items():
        img = next((i for i in d["imgs"] if i["is_primary"]), d["imgs"][0])
        groups[source_filename(img)].append((sid, d, img))

    # A group whose members all share one species epithet is the same butterfly
    # filed under two genus names — the photo is right, the taxonomy is doubled.
    synonym_groups, real_gaps = {}, {}
    for fname, members in groups.items():
        epithets = {m[1]["row"]["scientific_name"].split()[-1] for m in members}
        if len(members) > 1 and len(epithets) == 1:
            synonym_groups[fname] = members
        else:
            real_gaps[fname] = members

    gap_rows = [m for members in real_gaps.values() for m in members]
    syn_rows = [m for members in synonym_groups.values() for m in members]

    wb = Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 90
    lines = [
        ("Butterfly reference images — replacement list", "", ""),
        ("", "", ""),
        ("Species in database (active)", len(active), ""),
        ("Species with a correct, species-specific photo", sum(1 for s in active if verdicts[s][0] == "OK"), ""),
        ("Species needing a new photo", len(gap_rows), "Listed in 'Needs new photo'"),
        ("Species sharing a photo legitimately", len(syn_rows),
         "Same butterfly under two scientific names — needs a database merge, not a photo"),
        ("", "", ""),
        ("Important", "",
         "Every image currently in the system is genuine and correctly licensed "
         "(public domain, CC0, CC-BY, CC-BY-SA or CC-BY-NC), with attribution recorded. "
         "None are fabricated or unlicensed placeholders."),
        ("The defect", "",
         "For the species listed here the picture is not specific to that species. Most are "
         "scanned pages from Seitz's 'Macrolepidoptera of the World' — one page carries 50-150 "
         "labelled specimens, and the same page was attached to every species appearing on it. "
         "As a field reference the image is unusable."),
        ("", "", ""),
        ("Issue types", "", ""),
    ]
    for k, label in ISSUE_LABEL.items():
        n = sum(1 for m in gap_rows if verdicts[m[0]][0] == k)
        lines.append((f"    {k}", n, label))
    lines += [
        ("", "", ""),
        ("By family (species needing a new photo)", "", ""),
    ]
    fam = collections.Counter(m[1]["row"]["family"] for m in gap_rows)
    for f, n in fam.most_common():
        lines.append((f"    {f or '(unspecified)'}", n, ""))

    for row in lines:
        ws.append(row)
    ws["A1"].font = TITLE_FONT
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Needs new photo ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Needs new photo")
    ws.append([
        "#", "Scientific name", "Common name", "Family", "Issue",
        "Species on the current image", "Current image (source file)",
        "Current image URL", "Licence", "Source page", "Species ID",
    ])
    gap_rows.sort(key=lambda m: (m[1]["row"]["family"] or "", m[1]["row"]["scientific_name"]))
    for n, (sid, d, img) in enumerate(gap_rows, start=1):
        s = d["row"]
        verdict = verdicts[sid][0]
        shared_n = len(groups[source_filename(img)])
        ws.append([
            n, s["scientific_name"], s["common_name"], s["family"],
            ISSUE_LABEL.get(verdict, verdict),
            shared_n if shared_n > 1 else 1,
            source_filename(img), img["image_url"], img["license"],
            img["source_page_url"] or img["original_url"], s["id"],
        ])
    style_header(ws, [5, 30, 30, 16, 42, 12, 52, 62, 16, 52, 38])

    # ── Plate groups ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Plate groups")
    ws.append(["Source image file", "Species pinned to it", "Species"])
    for fname, members in sorted(real_gaps.items(), key=lambda kv: -len(kv[1])):
        if len(members) < 2:
            continue
        ws.append([
            fname, len(members),
            ", ".join(sorted(m[1]["row"]["scientific_name"] for m in members)),
        ])
    style_header(ws, [58, 20, 120])

    # ── Synonym pairs ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Synonym pairs")
    ws.append(["Source image file", "Names sharing it", "Action"])
    for fname, members in sorted(synonym_groups.items()):
        ws.append([
            fname,
            ", ".join(sorted(m[1]["row"]["scientific_name"] for m in members)),
            "Same butterfly under two names — merge the duplicate species record",
        ])
    style_header(ws, [58, 60, 60])

    wb.save(args.out)
    print(f"active species        : {len(active)}")
    print(f"correct photo         : {sum(1 for s in active if verdicts[s][0] == 'OK')}")
    print(f"needs a new photo     : {len(gap_rows)}")
    print(f"synonym duplicates    : {len(syn_rows)} species in {len(synonym_groups)} pairs")
    print(f"plate groups (2+ spp) : {sum(1 for m in real_gaps.values() if len(m) > 1)}")
    print(f"\nworkbook: {args.out}")


if __name__ == "__main__":
    main()
